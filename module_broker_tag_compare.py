import sqlalchemy as sa
from sqlalchemy import create_engine, text
import pymysql
import sqlite3

import numpy as np
import pandas as pd
from datetime import datetime, timedelta,date

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from data_logger import setup_logger
import module_query_update
import warnings
import os
warnings.filterwarnings('ignore')
warnings.filterwarnings('ignore', category=pd.errors.SettingWithCopyWarning)

logger = setup_logger()

def init(end_date=None):  
    istcok_db=getConnection('istock')
    rating_db=create_engine('sqlite:///result/rating_db.db')
    if end_date is None:
        end_date = pd.read_sql("SELECT MAX(td) FROM model_st ms ;",istcok_db)
        end_date=end_date.loc[0][0].strftime('%Y-%m-%d')
    start_date = pd.read_sql("select max(trade_date) from saas_rating;",rating_db)
    start_date=start_date.loc[0][0]
    sql_rating = f'''SELECT MIN(td_date) AS min_td_date
    FROM (
        SELECT  DISTINCT td_date 
        FROM shsz_stock_daily_quotation
        WHERE td_date <= '{end_date}'
        ORDER BY td_date DESC
        LIMIT 5
    ) AS recent_dates;'''
    date_df=pd.read_sql(sql_rating,istcok_db)
    from_date=date_df.loc[0][0].strftime('%Y-%m-%d')
    del date_df
    query_config= {'start_date': start_date,
                'end_date': end_date,
                'istock_conn': istcok_db,
                'rating_conn' : rating_db,
                'from_date': from_date}
    logger.info(f'同业对比初始化完成! 开始生成{end_date}的同业对比报告')
    return query_config


        
def get_rating(query_config,start_dates,end_dates):  
    sql_query = """
    SELECT trade_date, stock_code, wanlian_rating 
    FROM saas_rating 
    WHERE trade_date BETWEEN ? AND ?;
    """
    org_df=pd.read_sql(sql_query, query_config['rating_conn'], params=(start_dates, end_dates))

    from_date=pd.read_sql(f'''SELECT MAX(edit_date) FROM edit_log el  WHERE edit_date <='{start_dates}';''' ,query_config['rating_conn'])
    to_date=pd.read_sql(f'''SELECT MAX(edit_date) FROM edit_log el  WHERE edit_date <='{end_dates}';''' ,query_config['rating_conn'])

    def rating_edit_moudule(group,edit_log):
        if len(edit_log.loc[pd.to_datetime(edit_log['edit_date'])<=pd.to_datetime(group['trade_date'].unique()[0])])==0:
            group['new_rating'] = group['wanlian_rating']
            return group
        else:
            edit_log_date=edit_log.loc[pd.to_datetime(edit_log['edit_date'])<=pd.to_datetime(group['trade_date'].unique()[0])]
            edit_log_date=edit_log_date.loc[edit_log_date['edit_date']==edit_log_date['edit_date'].max()]
            group = pd.merge(group, edit_log_date[['stock_code', 'edit_rating']], on='stock_code', how='left')

            group['new_rating'] = group.apply(
                lambda row: row['edit_rating'] if pd.notnull(row['edit_rating']) else row['wanlian_rating'], axis=1)
            group.drop(columns=['edit_rating'], inplace=True)
            return group
        
    if (from_date.iloc[0].values[0] is None) and (to_date.iloc[0].values[0] is None):
        print('无白名单,机跑模型即为最终值')
        org_df.columns=['trade_date','stock_code','saas_five_class_result_adj']
    else:
        if from_date.iloc[0].values[0] is None:
            from_date=from_date.iloc[0].values[0]
            to_date=to_date.iloc[0].values[0]
            edit_log=pd.read_sql(f'''SELECT * FROM edit_log WHERE edit_date>='{to_date}' and edit_date<='{from_date}';''',query_config['rating_conn'])
            org_df=org_df.groupby('trade_date').apply(lambda x:rating_edit_moudule(x,edit_log)).reset_index(drop=True)
            org_df.columns=['trade_date','stock_code','backup_rating','saas_five_class_result_adj']
        else:
            from_date=from_date.iloc[0].values[0]
            to_date=to_date.iloc[0].values[0]
            edit_log=pd.read_sql(f'''SELECT * FROM edit_log WHERE edit_date>='{to_date}' and edit_date<='{from_date}';''',query_config['rating_conn'])
            org_df=org_df.groupby('trade_date').apply(lambda x:rating_edit_moudule(x,edit_log)).reset_index(drop=True)
            org_df.columns=['trade_date','stock_code','backup_rating','saas_five_class_result_adj']
    return org_df
        
def get_raw_data(query_config):
    rating_query = f'''SELECT r.stock_code, si.sec_short_name AS stock_name, r.trade_date, r.wanlian_rating
    FROM saas_rating r
    JOIN saas_indicator si 
        ON r.trade_date = si.trade_date 
        AND r.stock_code = si.stock_code
    WHERE r.trade_date BETWEEN '{query_config['from_date']}' AND '{query_config['end_date']}';'''
    saas_rating = pd.read_sql(rating_query, query_config['rating_conn'])

    sql_broker=f"""
            SELECT m.broker_name, m.stock_code , m.tra_date AS trade_date, m.concentration
                FROM margin_trading_records m
                INNER JOIN sec_basic_info s
                ON 
                    m.stock_code COLLATE utf8mb4_general_ci = s.sec_code COLLATE utf8mb4_general_ci
                    AND m.stock_name COLLATE utf8mb4_general_ci = s.sec_short_name_cn COLLATE utf8mb4_general_ci
                WHERE 
                    m.tra_date BETWEEN '{query_config['from_date']}' AND '{query_config['end_date']}'
                    AND m.broker_name IN ('中信建投',  '华泰证券', '国泰君安')
                    AND m.discount_rate IS NOT NULL
                    AND m.concentration REGEXP '^[^[:space:]]+$'
                    AND s.sec_type = 'A股'
                """
    rating_df=pd.read_sql(sql_broker,query_config['istock_conn'])
    condition = (rating_df['broker_name'] == '国泰君安') & (rating_df['concentration'].isin(['F', 'G']))
    rating_df.loc[condition, 'concentration'] = 'E'

    rating_df['trade_date'] = pd.to_datetime(rating_df['trade_date'])
    saas_rating['trade_date'] = pd.to_datetime(saas_rating['trade_date'])
    rating_sum=pd.merge(rating_df, saas_rating, on=['trade_date', 'stock_code'], how='left').dropna()
    rating_sum=rating_sum[['trade_date','stock_code','stock_name','broker_name','concentration','wanlian_rating']]       ###raw_data输出
    rating_sum.rename(columns={'wanlian_rating':'SAAS_rating'},inplace=True)
    saas_rating.rename(columns={'wanlian_rating':'SAAS_rating'},inplace=True)
    logger.info('同业对比raw_data准备完毕!')
    return rating_sum,saas_rating

def get_compare_data(query_config,saas_rating,raw_data):
    broker_compare = saas_rating[['stock_code', 'stock_name','trade_date', 'SAAS_rating']]
    for broker in list(raw_data['broker_name'].unique()):
        broker_compare = pd.merge(broker_compare, raw_data[raw_data['broker_name'] == broker][['stock_code','trade_date', 'concentration']],
                                on=['stock_code', 'trade_date'], how='left')
        broker_compare.rename(columns={'concentration': broker},inplace='True')

    def get_average_rating(group):
        rating_map = {'A': 5, 'B': 4, 'C': 3, 'D': 2, 'E': 1}
        reverse_rating_map = {v: k for k, v in rating_map.items()}
        group_numeric=group[['SAAS_rating','中信建投','华泰证券','国泰君安']].applymap(lambda x: rating_map.get(x, np.nan))
        avg_rating = group_numeric.mean()
        avg_ratings_mapped = avg_rating.apply(lambda x: reverse_rating_map.get(round(x)) if not np.isnan(x) else np.nan)
        return avg_ratings_mapped

    broker_compare_avg = broker_compare.groupby('stock_code').apply(get_average_rating).reset_index()
    broker_compare_avg.insert(1, 'trade_date', query_config['end_date'] )
    broker_compare['trade_date'] = pd.to_datetime(broker_compare['trade_date'], errors='coerce')
    broker_compare_avg['trade_date'] = pd.to_datetime(broker_compare_avg['trade_date'], errors='coerce')
    # 如果'stock_code'不是字符串类型，则转换
    broker_compare['stock_code'] = broker_compare['stock_code'].astype(str)
    broker_compare_avg['stock_code'] = broker_compare_avg['stock_code'].astype(str)
    broker_compare_avg = pd.merge(broker_compare_avg,broker_compare[['stock_code','trade_date', 'stock_name']],on=['stock_code','trade_date'],how='left')
    cols = list(broker_compare_avg.columns)  # 将现有列名转化为列表
    col_to_move = cols.pop(6)  # 移除第六列(索引为5)，因为pop默认移除并返回最后一个元素，所以这里直接给出索引
    cols.insert(1, col_to_move)
    broker_compare_avg = broker_compare_avg[cols]
    logger.info('整体数据集生成! 去重清洗完成!')
    return broker_compare,broker_compare_avg

def broker_compare_avg_adj(broker_compare,broker_compare_avg):
    broker_compares = broker_compare.copy()
    broker_compares.rename(columns={'SAAS_rating':'test'},inplace=True)
    broker_compare_avgs = broker_compare_avg.copy()
    broker_compare_avgs = pd.merge(broker_compare_avgs,broker_compares[['stock_code','trade_date', 'test']],
                                   on=['stock_code','trade_date'],how='left')
    broker_compare_avg['SAAS_rating']=broker_compare_avgs['test']
    return broker_compare_avg


def tag_maker_main(broker_compare_avg,query_config): 
    rating_map = {'A': 5, 'B': 4, 'C': 3, 'D': 2, 'E': 1}
    for i in ['SAAS_rating','中信建投','华泰证券','国泰君安']:
        broker_compare_avg[i+'_num']=broker_compare_avg[i].map(rating_map)
        broker_compare_avg['SAAS_rating_num'].astype(float)
    def tag_maker(row):
        try:
            broke_rating_num = row[['中信建投_num', '华泰证券_num', '国泰君安_num']]
            broke_rating = row[['中信建投', '华泰证券', '国泰君安']]
            stock_rating = row['SAAS_rating']
            stock_rating_num = row['SAAS_rating_num']
            if broke_rating.isnull().sum() == 0:
                if len(set(broke_rating))==1 :   # 三家券商一致的情况
                    if stock_rating_num - list(set(broke_rating_num))[0] >=2:
                        return 1
                    elif (list(set(broke_rating))[0] == 'A' or list(set(broke_rating))[0] == 'B') and stock_rating_num - list(set(broke_rating_num))[0] <= -2:
                        return 4
                    else:
                        return 5
                elif len(set(broke_rating))==2:    # 两家券商一致的情况
                    same_rating=broke_rating.value_counts()[broke_rating.value_counts()==2].index.tolist()[0]
                    same_rating_num=broke_rating_num.value_counts()[broke_rating_num.value_counts()==2].index.tolist()[0]
                    if same_rating == 'E' and stock_rating_num - same_rating_num >=2:
                        return 2
                    elif same_rating != 'E' and stock_rating_num - same_rating_num >=2:
                        return 3
                    elif (same_rating == 'A' or same_rating == 'B') and stock_rating_num - same_rating_num <=-2:
                        return 4
                    else:
                        return 5
                else:
                    return 5
            elif broke_rating.isnull().sum() == 1:
                row=row.dropna()
                broke_rating_num = row[[idx for idx in row.index if ('num' in idx) and ('SAAS' not in idx )]]
                broke_rating = row[[idx for idx in row.index if ('num' not in idx) and ('SAAS' not in idx ) and ('stock' not in idx ) and ('trade_date' not in idx )]]
                stock_rating = row['SAAS_rating']
                stock_rating_num = row['SAAS_rating_num']
                if len(set(broke_rating))==1 :   # 两家券商一致的情况
                    same_rating=broke_rating.value_counts()[broke_rating.value_counts()==2].index.tolist()[0]
                    same_rating_num=broke_rating_num.value_counts()[broke_rating_num.value_counts()==2].index.tolist()[0]
                    if same_rating == 'E' and stock_rating_num - same_rating_num >=2:
                        return 2
                    elif same_rating != 'E' and stock_rating_num - same_rating_num >=2:
                        return 3
                    elif (same_rating == 'A' or same_rating == 'B') and stock_rating_num - same_rating_num <=-2:
                        return 4
                    else:
                        return 5
                else:
                    return 5
            else:
                return 5
        except:
            logger.info(row)
    broker_compare_avg['tag']=broker_compare_avg.apply(tag_maker, axis=1)
    

    def tag_maker_1(row,yujing_df):
        stock_rating = row['SAAS_rating']
        stock_rating_num = row['SAAS_rating_num']
        tag_org = row['tag']
        if row['stock_code'] in yujing_df and tag_org==5 and stock_rating!='E':
            return 6
        else:
            return tag_org
        
    sql_yujing_query = f'''SELECT stock_code FROM indicator WHERE td='{query_config['end_date']}' AND ts_000 != 0; '''
    yujing_df=pd.read_sql(sql_yujing_query,query_config['istock_conn'])
    yujing_df['stock_code']=yujing_df['stock_code'].apply(lambda x:x[:6])
    yujing_df=yujing_df['stock_code'].tolist()
    broker_compare_avg['tag']=broker_compare_avg.apply(lambda x:tag_maker_1(x,yujing_df), axis=1)

    broker_compare_avg_tag=broker_compare_avg.copy()

    tag_dict=pd.DataFrame(tag_dict)
    logger.info('tag生成完成!')
    return broker_compare_avg_tag,tag_dict

def load_report_data(sheet_list,end_date):   
    with pd.ExcelWriter(f'report/{end_date}较同业券商偏离汇总报告.xlsx', engine='openpyxl') as writer:

        meau = pd.DataFrame(meau)
        def convert_trade_date_to_short_string(df):
            if not pd.api.types.is_datetime64_any_dtype(df['trade_date']):
                df['trade_date'] = pd.to_datetime(df['trade_date'])
            df['trade_date'] = df['trade_date'].dt.strftime('%Y-%m-%d')
            return df
        meau.to_excel(writer, sheet_name='目录', index=False)
        start_row = 0
        strat_col = len(meau.columns)+1
        sheet_list[4].to_excel(writer, sheet_name='目录', index=False, startrow=start_row, startcol=strat_col)
        sheet_list[0]=convert_trade_date_to_short_string(sheet_list[0])
        sheet_list[1]=convert_trade_date_to_short_string(sheet_list[1])
        sheet_list[2]=convert_trade_date_to_short_string(sheet_list[2])
        sheet_list[3]=convert_trade_date_to_short_string(sheet_list[3])
        sheet_list[0].to_excel(writer, sheet_name='rawdata', index=False)
        sheet_list[1].to_excel(writer, sheet_name='整体数据集', index=False)
        sheet_list[2].to_excel(writer, sheet_name='平滑后评级', index=False)
        bug_df=sheet_list[3].loc[sheet_list[3]['tag']!=5]
        bug_df['AI分析结果']=''
        bug_df['风险等级(AI实现)']=''
        bug_df['手工调整意见']=''
        start_row = 0
        strat_col = len(bug_df.columns)+2
        bug_df.to_excel(writer, sheet_name='异常票汇总', index=False, startrow=start_row, startcol=0)
        sheet_list[4].to_excel(writer, sheet_name='异常票汇总', index=False, startrow=start_row, startcol=strat_col)
        # module_query_update.insert_data_without_duplicates(sheet_list[3], 'broker_tag_sheet')    #在手动调整入库后再入rating_db

    def modify_excel(file_path, modifications):
        wb = load_workbook(filename=file_path)
        for sheet_name, column_width_dict in modifications.items():
            sheet = wb[sheet_name]
            # 如果 column_width_dict 为 None 或者为空，则设置所有列的宽度为15
            if not column_width_dict:
                max_col = sheet.max_column
                for col in range(1, max_col + 1):
                    col_letter = chr(64 + col)  # 将列索引转换为Excel列字母（例如：1 -> 'A'）
                    sheet.column_dimensions[col_letter].width = 15
            else:
                # 根据提供的列宽字典设置列宽
                for col_idx, width in column_width_dict.items():
                    sheet.column_dimensions[col_idx].width = width
        # 保存修改后的工作簿
        wb.save(file_path)
    # 定义要进行的修改
    modifications = 
    file_path = f'report/{end_date}较同业券商偏离汇总报告.xlsx'
    modify_excel(file_path, modifications)
    logger.info('报告已经生成,请查看report文件夹下文件,进行手动修改; 同业对比数据已经入库broker_tag_sheet!')


def broker_tag_compare(is_report=0):
    query_config=init()
    if pd.to_datetime(query_config['start_date'])<pd.to_datetime(query_config['end_date']):
        logger.info('需要取数,saas_rating取数更新中')
        module_query_update.daily_query_update()
    query_config = init()
    raw_data,saas_rating= get_raw_data(query_config)
    broker_compare,broker_compare_avg = get_compare_data(query_config,saas_rating,raw_data)
    broker_compare_avg = broker_compare_avg_adj(broker_compare,broker_compare_avg)
    broker_compare_avg_tag , tag_dict=tag_maker_main(broker_compare_avg.copy(),query_config)
    sheet_list=[raw_data,broker_compare,broker_compare_avg,broker_compare_avg_tag[['stock_code', 'stock_name', 'trade_date', 'SAAS_rating', '中信建投', '华泰证券',
        '国泰君安','tag']],tag_dict]
    
    if is_report==0:
        logger.info('不需要生成报告，直接入库!')
        sheet_list[3]['trade_date']=sheet_list[3]['trade_date'].dt.strftime('%Y-%m-%d')
        module_query_update.insert_data_without_duplicates(sheet_list[3], 'broker_tag_sheet')
    else:
        load_report_data(sheet_list,query_config['end_date'])
        logger.info('需要生成报告,请查看report文件夹下文件,进行手动修改; 同业对比数据已经入库broker_tag_sheet!')

def broker_compare_history(history_date):
    logger.info(f'{history_date}是上一次更新时间!  (n-1)')
    logger.info(f'以下是broker_tag_compare历史数据初始化!初始化日期不可信！！！ 开始取得上周历史修改数据!以{history_date}为准!')
    query_config=init(end_date=history_date)
    raw_data,saas_rating= get_raw_data(query_config)
    broker_compare,broker_compare_avg = get_compare_data(query_config,saas_rating,raw_data)
    broker_compare_avg = broker_compare_avg_adj(broker_compare,broker_compare_avg)
    broker_compare_avg_tag , tag_dict=tag_maker_main(broker_compare_avg.copy(),query_config)
    logger.info(f'{query_config["end_date"]}历史白名单数据已经回忆完成!')
    return broker_compare_avg_tag[['stock_code', 'stock_name', 'trade_date', 'SAAS_rating', 
                                   '中信建投', '华泰证券','国泰君安','tag']]


